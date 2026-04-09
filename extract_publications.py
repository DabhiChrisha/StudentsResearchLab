import openpyxl
import json

wb = openpyxl.load_workbook('d:\\StudentsResearchLab\\backend\\assets\\SRL Website - Data.xlsx')
print("Available sheets:", wb.sheetnames)

# Try to read all sheets
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"\n=== Sheet: {sheet_name} ===")
    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        if i <= 5:  # First 5 rows
            print(f"Row {i}: {row}")
        if i > 25:  # First 25 rows only
            break
